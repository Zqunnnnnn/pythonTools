import pymysql 
from sqlalchemy import create_engine
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import pandas as pd
import xlrd as xlrd
import seaborn as sns
import os 
os.chdir('C:\\Users\\sjkj\\Desktop\\代码和数据')

workbook = load_workbook('meal_order_detail.xlsx')
sheet_name = workbook.sheetnames
# print(sheet_name)
order1 = pd.read_excel('meal_order_detail.xlsx',sheet_name='meal_order_detail1')
order2 = pd.read_excel('meal_order_detail.xlsx',sheet_name='meal_order_detail2')
order3 = pd.read_excel('meal_order_detail.xlsx',sheet_name='meal_order_detail3')
#连接三表
order = pd.concat([order1,order2,order3],axis=0,ignore_index=False)#忽略原来的索引
print(order1.loc[104:105, 'dishes_name'][order1.loc[104:105, 'dishes_name'] == '焖猪手'])
