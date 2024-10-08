import requests
import openpyxl
import uuid

baseUrl = "https://api.bilibili.com/x/web-interface/popular?ps=20&pn=1&web_location=333.934&w_rid=44406e9d1e9a160e663059bc73d15c31&wts=1727248131"
header = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36"}


response = requests.get(url=baseUrl, headers=header)
list = response.json()['data']['list']
contents=[]
for item in list:
    content = {
        "title": item['title'],
        'desc':item['desc']
    }
    contents.append(content)

# print(contents)

workbook = openpyxl.Workbook()
sheet= workbook.active
sheet.title = 'b站热搜视频'
sheet.cell(row=1, column=1, value='b站热搜视频博主:Zqunnnnnn')

column = ("标题", "视频详情")  # Excel表头
for j in range(len(column)):
    sheet.cell(row=1, column=j+1, value=column[j])  # 使用j而不是j+1来正确索引

for i in range(len(contents)):
    sheet.cell(row=i+2, column=1, value=contents[i]['title'].strip())  # 使用 contents 列表中的每个元素
    sheet.cell(row=i+2, column=2, value=contents[i]['desc'].strip())  # 使用 contents 列表中的每个元素的 desc 值
# for i in range(len(contents)):
#         sheet.cell(row=i+2, column=1, value=contents[i]['title'].strip())#使用 for 循环遍历  列表中的每个元素，并使用 sheet.cell() 方法将每个元素写入工作表的相应单元格中。
#         sheet.cell(row=i+2, column=2, value=contents[i]['desc'])


# 生成一个UUID
unique_id = uuid.uuid4()

# 构建带有UUID的文件名
filename = f'saveFiles\\百度热搜{unique_id}.xlsx'
workbook.save(filename)#最后，使用 save() 方法将工作簿保存为一个名为 'baidu_hothit_list.xlsx' 的文件。
print('热搜数据已保存')
