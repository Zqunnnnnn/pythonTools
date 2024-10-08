import requests
from bs4 import BeautifulSoup
import openpyxl
import matplotlib.pyplot as plt
import uuid


url = 'https://top.baidu.com/board?tab=realtime'
response = requests.get(url)
html = response.content#这里使用了 requests.get() 方法发送GET请求，并将响应的内容赋值给变量 html。

soup = BeautifulSoup(html, 'html.parser')
hothit_list = []
#这段代码通过调用 soup.find_all() 方法找到所有 <div> 标签，并且指定 class 属性为 'c-single-text-ellipsis' 的元素。
#然后，将每个元素的文本内容添加到 hothit_list 列表中。
for item in soup.find_all('div', class_='c-single-text-ellipsis'):
    hothit_list.append(item.text.strip())#strip() 函数是用来清除字符串首尾指定的字符（如果不指定，则默认为空白字符）

#使用 openpyxl.Workbook() 创建一个新的工作簿对象。

#调用 active 属性获取当前活动的工作表对象，并将其赋值给变量 sheet。

#使用 title 属性给工作表命名为 'Baidu Hot Searches'。
workbook = openpyxl.Workbook()
sheet= workbook.active
sheet.title = '百度热搜'

#设置标题
sheet.cell(row=1, column=1, value='百度热搜排行榜—博主:Zqunnnnnn')

for i in range(len(hothit_list)):
    sheet.cell(row=i+2, column=1, value=hothit_list[i])#使用 for 循环遍历 hothit_list 列表中的每个元素，并使用 sheet.cell() 方法将每个元素写入工作表的相应单元格中。
# 生成一个UUID
unique_id = uuid.uuid4()

# 构建带有UUID的文件名
filename = f'saveFiles\\百度热搜{unique_id}.xlsx'
workbook.save(filename)#最后，使用 save() 方法将工作簿保存为一个名为 'baidu_hothit_list.xlsx' 的文件。
print('热搜数据已保存')


# # 设置中文字体
# plt.rcParams['font.sans-serif'] = ['SimHei']
# plt.rcParams['axes.unicode_minus'] = False
 
# # 绘制条形图
# plt.figure(figsize=(15, 10))
# x = range(len(hothit_list))
# y = list(reversed(range(1, len(hothit_list)+1)))
# plt.barh(x, y, tick_label=hothit_list, height=0.8)  # 调整条形图的高度
 
# # 添加标题和标签
# plt.title('百度热搜排行榜')
# plt.xlabel('排名')
# plt.ylabel('关键词')
 
# # 调整坐标轴刻度
# plt.xticks(range(1, len(hothit_list)+1))
 
# # 调整条形图之间的间隔
# plt.subplots_adjust(hspace=0.8, wspace=0.5)
 
# # 显示图形
# plt.tight_layout()
# plt.show()
